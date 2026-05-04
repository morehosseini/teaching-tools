"""
Excel workbook export — 8-tab professional planning package (§20).

Generates a formatted Excel workbook with:
1. Executive Summary
2. Project Information Request
3. Basis of Schedule
4. WBS and Dictionary
5. Full Schedule
6. Procurement Schedule
7. Assumptions and Risk Register
8. Validation Warnings and Sign-Off
"""

from __future__ import annotations

import datetime
import io
from typing import Optional

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side, numbers
)
from openpyxl.utils import get_column_letter

from .models import Project, Activity, WBSElement, ProcurementItem, ValidationResult, Severity


# ── Styles ─────────────────────────────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="2F5496")
SUBTITLE_FONT = Font(name="Calibri", size=11, bold=True, color="2F5496")
NORMAL_FONT = Font(name="Calibri", size=10)
CRITICAL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ERROR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
OK_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _apply_header_style(ws, row: int, cols: int):
    """Apply header styling to a row."""
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _auto_width(ws, min_width: int = 10, max_width: int = 40):
    """Auto-adjust column widths."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                val = str(cell.value) if cell.value else ""
                max_len = max(max_len, len(val))
            except Exception:
                pass
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


# ── Tab 1: Executive Summary ──────────────────────────────────────────────────

def _create_executive_summary(wb: openpyxl.Workbook, project: Project):
    ws = wb.active
    ws.title = "Executive Summary"

    # Title
    ws["A1"] = "CONSTRUCTION PLANNING AGENT — EXECUTIVE SUMMARY"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    # Project info
    rows = [
        ("Project Name", project.project_name),
        ("Project ID", project.project_id),
        ("Generated", datetime.datetime.now().strftime("%d %B %Y, %H:%M")),
        ("AACE Estimate Class", project.planning_basis.aace_class.value if project.planning_basis else "Class 5"),
        ("", ""),
        ("PROJECT COMPLETION", ""),
        ("P50 Completion", str(project.p50_completion) if project.p50_completion else "Not calculated"),
        ("P80 Completion", str(project.p80_completion) if project.p80_completion else "Not calculated"),
        ("Project Start", str(project.project_start_date)),
        ("Schedule Version", project.schedule_version),
        ("", ""),
        ("SCHEDULE SUMMARY", ""),
        ("Total Activities", len(project.activities)),
        ("WBS Elements", len(project.wbs_elements)),
        ("Procurement Items", len(project.procurement_items)),
        ("Validation Errors", len([v for v in project.validation_results if v.severity == Severity.ERROR])),
        ("Validation Warnings", len([v for v in project.validation_results if v.severity == Severity.WARNING])),
        ("", ""),
        ("DISCLAIMER", ""),
    ]

    for i, (label, value) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=label).font = SUBTITLE_FONT if label.isupper() else NORMAL_FONT
        ws.cell(row=i, column=2, value=value).font = NORMAL_FONT

    # Disclaimer
    disclaimer_row = len(rows) + 4
    ws.cell(row=disclaimer_row, column=1, value=project.disclaimer).font = Font(
        name="Calibri", size=9, italic=True, color="808080"
    )
    ws.merge_cells(f"A{disclaimer_row}:D{disclaimer_row + 2}")
    ws.cell(row=disclaimer_row, column=1).alignment = Alignment(wrap_text=True)

    _auto_width(ws)


# ── Tab 2: PIR ─────────────────────────────────────────────────────────────────

def _create_pir_tab(wb: openpyxl.Workbook, project: Project):
    ws = wb.create_sheet("Project Information")

    headers = ["Variable", "Value", "Status", "Default", "Rationale", "Schedule Impact", "Source"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _apply_header_style(ws, 1, len(headers))

    if project.planning_basis:
        for i, pir in enumerate(project.planning_basis.pir_variables, start=2):
            ws.cell(row=i, column=1, value=pir.label).font = NORMAL_FONT
            ws.cell(row=i, column=2, value=str(pir.value) if pir.value else "").font = NORMAL_FONT
            ws.cell(row=i, column=3, value=pir.status.value).font = NORMAL_FONT
            ws.cell(row=i, column=4, value=str(pir.default_value) if pir.default_value else "").font = NORMAL_FONT
            ws.cell(row=i, column=5, value=pir.default_rationale or "").font = NORMAL_FONT
            ws.cell(row=i, column=6, value=pir.schedule_impact or "").font = NORMAL_FONT
            ws.cell(row=i, column=7, value=pir.source or "").font = NORMAL_FONT

    _auto_width(ws)


# ── Tab 3: Basis of Schedule ──────────────────────────────────────────────────

def _create_basis_tab(wb: openpyxl.Workbook, project: Project):
    ws = wb.create_sheet("Basis of Schedule")
    ws["A1"] = "BASIS OF SCHEDULE"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    narrative = project.basis_of_schedule_narrative or "Basis of Schedule not yet generated."
    ws.cell(row=3, column=1, value=narrative).font = NORMAL_FONT
    ws.cell(row=3, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells("A3:D20")
    ws.column_dimensions["A"].width = 80


# ── Tab 4: WBS Dictionary ─────────────────────────────────────────────────────

def _create_wbs_tab(wb: openpyxl.Workbook, project: Project):
    ws = wb.create_sheet("WBS Dictionary")

    headers = ["WBS Code", "Parent", "Name", "Level", "Description", "Confidence"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _apply_header_style(ws, 1, len(headers))

    for i, elem in enumerate(project.wbs_elements, start=2):
        ws.cell(row=i, column=1, value=elem.wbs_code).font = NORMAL_FONT
        ws.cell(row=i, column=2, value=elem.parent_code or "").font = NORMAL_FONT
        indent = "  " * (elem.level - 1)
        ws.cell(row=i, column=3, value=f"{indent}{elem.name}").font = NORMAL_FONT
        ws.cell(row=i, column=4, value=elem.level).font = NORMAL_FONT
        ws.cell(row=i, column=5, value=elem.description or "").font = NORMAL_FONT
        ws.cell(row=i, column=6, value=elem.confidence_level.value).font = NORMAL_FONT

    _auto_width(ws)


# ── Tab 5: Full Schedule ──────────────────────────────────────────────────────

def _create_schedule_tab(wb: openpyxl.Workbook, project: Project):
    ws = wb.create_sheet("Full Schedule")

    headers = [
        "ID", "WBS", "Activity Name", "Type", "Trade", "Location",
        "Qty", "Unit", "Rate Source",
        "Opt (d)", "ML (d)", "Pess (d)",
        "Calendar", "Predecessors",
        "ES", "EF", "LS", "LF", "TF", "FF", "Critical",
        "Confidence", "Review Required", "Assumption",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _apply_header_style(ws, 1, len(headers))

    for i, act in enumerate(project.activities, start=2):
        row_data = [
            act.activity_id,
            act.wbs_code,
            act.activity_name,
            act.activity_type.value,
            act.trade or "",
            act.location_zone or "",
            act.quantity,
            act.unit or "",
            act.production_rate_source or "",
            act.duration_optimistic_days,
            act.duration_most_likely_days,
            act.duration_pessimistic_days,
            act.calendar_id,
            ", ".join(f"{p.activity_id} {p.relationship_type.value}+{p.lag_days}d" for p in act.predecessors),
            str(act.early_start) if act.early_start else "",
            str(act.early_finish) if act.early_finish else "",
            str(act.late_start) if act.late_start else "",
            str(act.late_finish) if act.late_finish else "",
            act.total_float,
            act.free_float,
            "YES" if act.is_critical else "",
            act.confidence_level.value,
            "YES" if act.human_review_required else "",
            act.assumption or "",
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            if act.is_critical:
                cell.fill = CRITICAL_FILL

    ws.freeze_panes = "D2"
    _auto_width(ws)


# ── Tab 6: Procurement ────────────────────────────────────────────────────────

def _create_procurement_tab(wb: openpyxl.Workbook, project: Project):
    ws = wb.create_sheet("Procurement")

    headers = [
        "Item ID", "Category", "Description",
        "Design Freeze (d)", "Shop Drawings (d)", "Consultant Review (d)",
        "Approval (d)", "Fabrication (d)", "Delivery (d)",
        "Lead Min (wk)", "Lead Max (wk)",
        "Installation Activity", "Source", "Notes",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _apply_header_style(ws, 1, len(headers))

    for i, item in enumerate(project.procurement_items, start=2):
        row_data = [
            item.item_id,
            item.item_category,
            item.description,
            item.design_freeze_days,
            item.shop_drawing_days,
            item.consultant_review_days,
            item.approval_days,
            item.fabrication_days,
            item.delivery_days,
            item.total_lead_weeks_min,
            item.total_lead_weeks_max,
            item.installation_activity_id or "",
            item.source or "",
            item.notes or "",
        ]
        for col, val in enumerate(row_data, 1):
            ws.cell(row=i, column=col, value=val).font = NORMAL_FONT

    _auto_width(ws)


# ── Tab 7: Assumptions & Risk Register ────────────────────────────────────────

def _create_risk_tab(wb: openpyxl.Workbook, project: Project):
    ws = wb.create_sheet("Assumptions & Risks")

    # Assumptions section
    ws["A1"] = "ASSUMPTIONS"
    ws["A1"].font = TITLE_FONT
    if project.planning_basis:
        for i, assumption in enumerate(project.planning_basis.assumptions, start=2):
            ws.cell(row=i, column=1, value=f"• {assumption}").font = NORMAL_FONT

    # Risk register
    risk_start = max(3, len(project.planning_basis.assumptions) + 3 if project.planning_basis else 3)
    ws.cell(row=risk_start, column=1, value="RISK REGISTER").font = TITLE_FONT

    headers = ["Risk ID", "Description", "Category", "Likelihood", "Impact", "Linked Activities", "Mitigation", "Owner", "Status"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=risk_start + 1, column=col, value=header)
    _apply_header_style(ws, risk_start + 1, len(headers))

    for i, risk in enumerate(project.risks, start=risk_start + 2):
        row_data = [
            risk.risk_id,
            risk.description,
            risk.category,
            risk.likelihood,
            risk.impact,
            ", ".join(risk.linked_activities),
            risk.mitigation or "",
            risk.owner or "",
            risk.status,
        ]
        for col, val in enumerate(row_data, 1):
            ws.cell(row=i, column=col, value=val).font = NORMAL_FONT

    _auto_width(ws)


# ── Tab 8: Validation & Sign-Off ──────────────────────────────────────────────

def _create_validation_tab(wb: openpyxl.Workbook, project: Project):
    ws = wb.create_sheet("Validation & Sign-Off")

    # Validation results
    ws["A1"] = "VALIDATION RESULTS"
    ws["A1"].font = TITLE_FONT

    headers = ["Rule ID", "Severity", "Description", "Affected Activities", "Suggested Fix", "Source", "Overridden"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=header)
    _apply_header_style(ws, 2, len(headers))

    for i, vr in enumerate(project.validation_results, start=3):
        row_data = [
            vr.rule_id,
            vr.severity.value,
            vr.description,
            ", ".join(vr.affected_activities),
            vr.suggested_fix or "",
            vr.source or "",
            "YES" if vr.overridden else "",
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = NORMAL_FONT
            if vr.severity == Severity.ERROR:
                cell.fill = ERROR_FILL
            elif vr.severity == Severity.WARNING:
                cell.fill = WARNING_FILL

    # Sign-Off section
    signoff_start = max(5, len(project.validation_results) + 5)
    ws.cell(row=signoff_start, column=1, value="PLANNER SIGN-OFF").font = TITLE_FONT

    checklist = [
        "All high-uncertainty items reviewed and confirmed or adjusted",
        "Procurement lead times cross-checked with current supplier quotes",
        "Weather and calendar assumptions accepted for the project location and season",
        "Validation warnings reviewed and either accepted or mitigated",
        "Basis of Schedule narrative reviewed for accuracy",
    ]

    for i, item in enumerate(checklist, start=signoff_start + 1):
        ws.cell(row=i, column=1, value=f"☐  {item}").font = NORMAL_FONT
        ws.cell(row=i, column=5, value="").font = NORMAL_FONT  # Checkbox column

    sig_row = signoff_start + len(checklist) + 2
    ws.cell(row=sig_row, column=1, value="Planner Name:").font = SUBTITLE_FONT
    ws.cell(row=sig_row + 1, column=1, value="Professional Registration:").font = SUBTITLE_FONT
    ws.cell(row=sig_row + 2, column=1, value="Date:").font = SUBTITLE_FONT
    ws.cell(row=sig_row + 3, column=1, value="Signature:").font = SUBTITLE_FONT

    _auto_width(ws)


# ── Main export function ──────────────────────────────────────────────────────

def export_project_to_excel(project: Project) -> bytes:
    """
    Export the full project to an 8-tab Excel workbook.
    Returns the workbook as bytes for download.
    """
    wb = openpyxl.Workbook()

    _create_executive_summary(wb, project)
    _create_pir_tab(wb, project)
    _create_basis_tab(wb, project)
    _create_wbs_tab(wb, project)
    _create_schedule_tab(wb, project)
    _create_procurement_tab(wb, project)
    _create_risk_tab(wb, project)
    _create_validation_tab(wb, project)

    # Write to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
